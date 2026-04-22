from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from datetime import datetime
import io
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer, Paragraph
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER
from app.models.database import SessionLocal
from app.models.user import User
from app.models.task import UploadTask
from app.routes.auth import get_current_user

router = APIRouter(prefix="/export", tags=["export"])

USD_TO_INR = 93.78

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

@router.get("/excel")
def export_excel(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Export all user data to Excel"""
    tasks = db.query(UploadTask).filter(
        UploadTask.user_id == current_user.id
    ).order_by(UploadTask.created_at.desc()).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Waste Analysis Report"
    
    # Header
    ws['A1'] = "Waste LCA AI - Analysis Report"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A4'] = f"User: {current_user.email}"
    
    # Headers
    headers = ["ID", "File Name", "Status", "Upload Date", "Materials"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    for row, task in enumerate(tasks, 7):
        materials_str = ""
        if task.extracted_materials:
            try:
                materials = json.loads(task.extracted_materials)
                materials_str = ", ".join([f"{m['name']}({m['percentage']}%)" for m in materials])
            except:
                pass
        
        ws.cell(row=row, column=1, value=task.id)
        ws.cell(row=row, column=2, value=task.file_name)
        ws.cell(row=row, column=3, value=task.status)
        ws.cell(row=row, column=4, value=task.created_at.strftime("%Y-%m-%d %H:%M") if task.created_at else "")
        ws.cell(row=row, column=5, value=materials_str)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=waste_lca_report_{datetime.now().strftime('%Y%m%d')}.xlsx"}
    )

@router.get("/pdf")
def export_pdf(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Export all user data to PDF"""
    tasks = db.query(UploadTask).filter(
        UploadTask.user_id == current_user.id
    ).order_by(UploadTask.created_at.desc()).all()
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    
    story = []
    story.append(Paragraph("Waste LCA AI - Analysis Report", styles['Title']))
    story.append(Spacer(1, 0.2*inch))
    story.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
    story.append(Paragraph(f"User: {current_user.email}", styles['Normal']))
    story.append(Spacer(1, 0.3*inch))
    
    data = [["ID", "File Name", "Status", "Date"]]
    for task in tasks[:50]:
        data.append([
            str(task.id),
            task.file_name[:40],
            task.status,
            task.created_at.strftime("%Y-%m-%d") if task.created_at else ""
        ])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    story.append(table)
    
    doc.build(story)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=waste_lca_report_{datetime.now().strftime('%Y%m%d')}.pdf"}
    )